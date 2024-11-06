from reportlab.lib.pagesizes import letter  # reportlabからページサイズをインポート
from reportlab.pdfgen import canvas  # reportlabからcanvasをインポート
from openpyxl import load_workbook  # Excelファイルを読み込むためにload_workbookをインポート

# 既存のExcelファイルのパス
file_path = '/Users/tenhou/Desktop/test.xlsx'  # Mac用のパスに変更

# Excelファイルを開く
workbook = load_workbook(file_path)  # 既存のExcelファイルを読み込む
worksheet = workbook.active  # 最初のワークシートを取得

# PDFとして保存
pdf_path = '/Users/tenhou/Desktop/output.pdf'  # 保存先をDesktopに設定

# PDFを作成
def create_pdf(pdf_path):
    c = canvas.Canvas(pdf_path, pagesize=letter)  # PDFキャンバスを作成
    # セルの値をPDFに書き込む
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            c.drawString(100, 750 - (cell.row - 1) * 20, f"{cell.value}")  # セルの値をPDFに追加
    c.save()  # PDFを保存

# PDFとして保存
create_pdf(pdf_path)  # PDFを作成